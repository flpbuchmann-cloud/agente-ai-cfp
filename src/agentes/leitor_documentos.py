"""
Leitor de documentos para os agentes.

Suporta: PDF, DOCX, XLSX, TXT, CSV, imagens (PNG, JPG).
Extrai texto e retorna conteúdo formatado para os agentes.
"""

import os
import base64
from pathlib import Path


def ler_pdf(caminho: str) -> str:
    """Extrai texto de um PDF."""
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(caminho)
        textos = []
        for i, page in enumerate(reader.pages):
            texto = page.extract_text()
            if texto and texto.strip():
                textos.append(f"--- Página {i+1} ---\n{texto.strip()}")
        if textos:
            return "\n\n".join(textos)
        return "[PDF sem texto extraível — pode ser imagem/escaneado]"
    except Exception as e:
        return f"[Erro ao ler PDF: {e}]"


def ler_docx(caminho: str) -> str:
    """Extrai texto de um DOCX."""
    try:
        from docx import Document
        doc = Document(caminho)
        paragrafos = [p.text for p in doc.paragraphs if p.text.strip()]
        if paragrafos:
            return "\n".join(paragrafos)
        return "[DOCX vazio]"
    except Exception as e:
        return f"[Erro ao ler DOCX: {e}]"


def ler_xlsx(caminho: str) -> str:
    """Extrai texto de um XLSX (todas as sheets)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(caminho, data_only=True)
        resultado = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            linhas = []
            for row in ws.iter_rows(values_only=True):
                valores = [str(c) if c is not None else "" for c in row]
                if any(v.strip() for v in valores):
                    linhas.append(" | ".join(valores))
            if linhas:
                resultado.append(f"=== Planilha: {sheet_name} ===\n" + "\n".join(linhas))
        if resultado:
            return "\n\n".join(resultado)
        return "[XLSX vazio]"
    except Exception as e:
        return f"[Erro ao ler XLSX: {e}]"


def ler_texto(caminho: str) -> str:
    """Lê arquivo de texto simples (TXT, CSV)."""
    try:
        with open(caminho, "r", encoding="utf-8") as f:
            return f.read()
    except UnicodeDecodeError:
        with open(caminho, "r", encoding="latin-1") as f:
            return f.read()
    except Exception as e:
        return f"[Erro ao ler texto: {e}]"


def imagem_para_base64(caminho: str) -> dict:
    """Converte imagem para base64 para envio via API Claude (vision)."""
    ext = Path(caminho).suffix.lower()
    media_types = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".webp": "image/webp",
    }
    media_type = media_types.get(ext, "image/png")

    with open(caminho, "rb") as f:
        data = base64.standard_b64encode(f.read()).decode("utf-8")

    return {
        "type": "image",
        "source": {
            "type": "base64",
            "media_type": media_type,
            "data": data,
        },
    }


def ler_documento(caminho: str) -> dict:
    """
    Lê um documento e retorna seu conteúdo.

    Returns:
        {
            "nome": "arquivo.pdf",
            "tipo": "pdf",
            "texto": "conteúdo extraído...",
            "imagem": None ou dict base64 (para imagens)
        }
    """
    nome = os.path.basename(caminho)
    ext = Path(caminho).suffix.lower()

    resultado = {
        "nome": nome,
        "tipo": ext.lstrip("."),
        "texto": None,
        "imagem": None,
    }

    if ext == ".pdf":
        resultado["texto"] = ler_pdf(caminho)
    elif ext in (".docx", ".doc"):
        resultado["texto"] = ler_docx(caminho)
    elif ext in (".xlsx", ".xls"):
        resultado["texto"] = ler_xlsx(caminho)
    elif ext in (".txt", ".csv", ".json", ".md"):
        resultado["texto"] = ler_texto(caminho)
    elif ext in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
        resultado["imagem"] = imagem_para_base64(caminho)
        resultado["texto"] = f"[Imagem: {nome}]"
    else:
        resultado["texto"] = f"[Formato não suportado: {ext}]"

    return resultado


def ler_pasta(pasta: str) -> list[dict]:
    """Lê todos os documentos de uma pasta."""
    if not os.path.isdir(pasta):
        return []

    documentos = []
    extensoes_validas = {
        ".pdf", ".docx", ".doc", ".xlsx", ".xls",
        ".txt", ".csv", ".json", ".md",
        ".png", ".jpg", ".jpeg", ".gif", ".webp",
    }

    for arquivo in sorted(os.listdir(pasta)):
        caminho = os.path.join(pasta, arquivo)
        if os.path.isfile(caminho) and Path(arquivo).suffix.lower() in extensoes_validas:
            documentos.append(ler_documento(caminho))

    return documentos


def formatar_documentos_para_prompt(documentos: list[dict]) -> str:
    """Formata lista de documentos para inclusão no prompt do agente."""
    if not documentos:
        return "[NENHUM DOCUMENTO DISPONÍVEL NESTA ÁREA]\n\nSinalize no relatório quais documentos seriam necessários para uma análise completa."

    partes = []
    for doc in documentos:
        if doc["texto"]:
            partes.append(f"### Documento: {doc['nome']}\n{doc['texto']}")

    return "\n\n---\n\n".join(partes)
